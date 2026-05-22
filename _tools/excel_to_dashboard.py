"""
Excel → Dashboard HTML converter (Two Build / One Vela project)

Reads the daily Excel export "การขายลงทุกแปลงทั้งหมด" from Ecount ERP
and produces an updated index.html with fresh cost data.

Usage:
    python3 excel_to_dashboard.py <path-to-excel.xlsx>

Output:
    Writes new index.html into the dashboard-online folder.
    Also writes deploy zip ready for Netlify drag-drop.

Architecture:
    - Reads plot metadata (483 plots) from existing index.html
    - Reads item→category mapping from existing item_categories.json
    - Reads Excel transactions, aggregates per plot
    - Merges Excel costs with plot metadata
    - Re-injects PROJECT_DATA JSON into index.html template
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from zipfile import ZipFile, ZIP_DEFLATED

# ───────────────────────────────────────────────
# Configuration
# ───────────────────────────────────────────────
DASHBOARD_DIR = Path(__file__).parent.parent
TOOLS_DIR = Path(__file__).parent
TEMPLATE_HTML = DASHBOARD_DIR / 'index.html'       # current deployed HTML — used as template
METADATA_JSON = TOOLS_DIR / 'plots_metadata.json'  # 483 plots metadata
CATEGORIES_JSON = TOOLS_DIR / 'item_categories.json'  # item name → category dictionary
OUT_HTML = DASHBOARD_DIR / 'index.html'
OUT_ZIP = DASHBOARD_DIR / 'dashboard-deploy.zip'
OUT_REPORT_DIR = DASHBOARD_DIR / '_audit_reports'

# Phase 2 — Google Sheet metadata sync
# Sheet URL is read from sheet_config.json (created when you set up the Sheet).
SHEET_CONFIG_PATH = TOOLS_DIR / 'sheet_config.json'

# Phase → Category mapping (8 phases ← 8 specific categories;
# ค่าแรง and วัสดุสิ้นเปลือง/อื่นๆ are tracked but not assigned to a phase)
PHASE_CATEGORIES = [
    ('โครงสร้าง',       ['งานโครงสร้าง']),
    ('ก่อ-ฉาบผนัง',     ['งานก่อ-ฉาบ ผนัง']),
    ('หลังคา',          ['งานหลังคา']),
    ('ประตู-หน้าต่าง',   ['งานประตู-หน้าต่าง']),
    ('พื้น-กระเบื้อง',   ['งานพื้น-กระเบื้องบุ']),
    ('งานระบบ',         ['งานระบบไฟฟ้า', 'งานระบบประปา/กันซึม']),
    ('สี',              ['งานสี']),
    ('สุขภัณฑ์',         ['งานสุขภัณฑ์']),
]

# Excel plot-type string → Dashboard typeKey
TYPE_KEY_MAP = {
    'บ้านเดี่ยว 2 ชั้น':   'A',
    'บ้านเดี่ยวชั้นเดียว':  'B',
    'บ้านแฝดชั้นเดียว':    'C',
    'บ้านทาวน์โฮม':       'D',
    'ทาวน์โฮม':           'D',
}

# Thai month abbreviations (in date strings "06 เม.ย. 2569")
THAI_MONTHS = {
    'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4,
    'พ.ค.': 5, 'มิ.ย.': 6, 'ก.ค.': 7, 'ส.ค.': 8,
    'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12,
}

# Manual category overrides for items not yet in the dictionary
MANUAL_CATEGORIES = {
    'ช่างแอ็ด-แม็กโคร-หกล้อบรรทุก': 'ค่าแรง',
    'สังกะสี 12 F':           'งานหลังคา',
    'สังกะสี 7F':              'งานหลังคา',
    'ไม้ยูคา 2.5นิ้ว*3นิ้ว':    'งานโครงสร้าง',
    'ไม้ยูคา 2 นิ้ว คูณ 3 นิ้ว': 'งานโครงสร้าง',
    'ยางนก 3*3':               'วัสดุสิ้นเปลือง/อื่นๆ',
    'อัดดำ 15 มล.':            'วัสดุสิ้นเปลือง/อื่นๆ',
}


# ───────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────
def parse_thai_date(s):
    """'06 เม.ย. 2569 ' → ('2026-04-06', '6 เม.ย. 69')"""
    s = s.strip()
    m = re.match(r'(\d{1,2})\s+([ก-๙.]+)\s+(\d{4})', s)
    if not m:
        return None, s
    day = int(m.group(1))
    month_th = m.group(2)
    year_th = int(m.group(3))
    month = THAI_MONTHS.get(month_th)
    if not month:
        return None, s
    year_ce = year_th - 543
    iso = f'{year_ce:04d}-{month:02d}-{day:02d}'
    short = f'{day} {month_th} {year_th % 100:02d}'
    return iso, short


def classify_item(name, item_dict):
    """Return category for a material/labor item name."""
    if name in item_dict:
        return item_dict[name]
    if name in MANUAL_CATEGORIES:
        return MANUAL_CATEGORIES[name]
    # Fallback: starts with ช่าง = labor
    if name.startswith('ช่าง') or 'ค่าบริการ' in name:
        return 'ค่าแรง'
    # Unknown → consumables
    return 'วัสดุสิ้นเปลือง/อื่นๆ'


def round2(x):
    return round(float(x or 0) + 1e-9, 2)


def _cell(row, idx):
    """อ่านค่าเซลล์อย่างปลอดภัย — กัน index เกินความยาวแถว"""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def detect_columns(header_row):
    """
    ตรวจหาตำแหน่งคอลัมน์จากแถวหัวตาราง — รองรับไฟล์ Ecount หลายรูปแบบ
    เช่น  รูปแบบเดิม : ประจำวัน | ลูกค้า/ผู้ขาย | ตามสินค้า | จำนวน | จำนวนเงิน | ภาษี | ยอดรวม
         รูปแบบใหม่ : วันที่-ลำดับ | ชื่อสินค้า | ชื่อลูกค้า/ผู้ขาย | จำนวน | ราคาต่อหน่วย | จำนวนเงิน | ภาษี | ยอดรวม
    คืน dict {date, plot, item, qty, amount, total} → ดัชนีคอลัมน์ (0-based)
    ถ้าตรวจไม่พบ ใช้ค่าเริ่มต้น = รูปแบบเดิม
    """
    cols = {'date': 0, 'plot': 1, 'item': 2, 'qty': 3, 'amount': 4, 'total': 6}
    if not header_row:
        return cols
    found = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        t = str(h).strip()
        if not t:
            continue
        if 'วัน' in t and 'date' not in found:
            found['date'] = i
        elif 'สินค้า' in t and 'item' not in found:
            found['item'] = i
        elif ('ลูกค้า' in t or 'ผู้ขาย' in t) and 'plot' not in found:
            found['plot'] = i
        elif t == 'จำนวน' and 'qty' not in found:
            found['qty'] = i
        elif 'จำนวนเงิน' in t and 'amount' not in found:
            found['amount'] = i
        elif 'ยอดรวม' in t and 'total' not in found:
            found['total'] = i
    cols.update(found)
    return cols


# ───────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────
def convert(excel_path: Path):
    print(f"\n📊 Converting: {excel_path.name}")

    # ─── Load resources ───
    meta = json.loads(METADATA_JSON.read_text(encoding='utf-8'))
    plot_metadata = {p['plot']: p for p in meta['houses_metadata']}
    item_dict = json.loads(CATEGORIES_JSON.read_text(encoding='utf-8'))
    print(f"   Loaded {len(plot_metadata)} plots metadata, {len(item_dict)} item mappings")

    # ─── Read Excel ───
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    print(f"   Excel: '{ws.title}', {len(rows)} rows")

    # ─── ตรวจหาคอลัมน์จากหัวตาราง (รองรับ Ecount หลายรูปแบบ) ───
    C = detect_columns(rows[1] if len(rows) > 1 else None)
    print(f"   คอลัมน์ที่ใช้: วันที่=#{C['date']} แปลง=#{C['plot']} สินค้า=#{C['item']} "
          f"จำนวน=#{C['qty']} ยอดรวม=#{C['total']}")

    # ─── Group transactions by plot — with audit trail ───
    plot_txns = defaultdict(list)
    excluded_rows = []   # (row_num, reason, A, B, C, qty, amount, total)
    new_items = set()    # items not in dictionary
    all_input_rows = []  # for dup detection — every detail row seen
    excel_total_from_summary = 0  # cross-check: sum of per-plot summary rows ONLY (avoids double counting)

    # Per-plot summary regex (e.g., "บ้านเดี่ยว 2 ชั้น แปลง 3 ทั้งหมด")
    per_plot_sum_re = re.compile(r'แปลง\s*\d+\s*ทั้งหมด')
    # Footer timestamp pattern at the bottom of Ecount exports
    footer_re = re.compile(r'^\d{1,2}\s+[ก-๙.]+\s+\d{4}.*\d{1,2}:\d{2}')

    for excel_row_idx, row in enumerate(rows[2:], start=3):  # excel row index (1-based, accounting for header)
        va = _cell(row, C['date']); col_a = (str(va) or '').strip() if va else ''
        vb = _cell(row, C['plot']); col_b = (str(vb) or '').strip() if vb else ''
        vc = _cell(row, C['item']); col_c = (str(vc) or '').strip() if vc else ''

        # Summary row "...ทั้งหมด" — only count per-plot summaries for cross-check
        if 'ทั้งหมด' in col_a:
            if per_plot_sum_re.search(col_a):
                excel_total_from_summary += round2(_cell(row, C['total']))
            continue
        # Footer timestamp row at end of Ecount export — skip silently
        if col_a and not col_b and not col_c and footer_re.match(col_a):
            continue
        # Empty/incomplete row
        if not col_a or not col_b or not col_c:
            if any([col_a, col_b, col_c]):
                excluded_rows.append((excel_row_idx, 'ข้อมูลไม่ครบ', col_a, col_b, col_c, _cell(row, C['qty']), _cell(row, C['amount']), _cell(row, C['total'])))
            continue
        # Parse plot identifier
        m = re.match(r'^(.+?)\s*แปลง\s*(\d+)', col_b)
        if not m:
            excluded_rows.append((excel_row_idx, 'ไม่ใช่รายการของแปลง (overhead)', col_a, col_b, col_c, _cell(row, C['qty']), _cell(row, C['amount']), _cell(row, C['total'])))
            continue
        type_str = m.group(1).strip()
        plot_num = int(m.group(2))
        if plot_num not in plot_metadata:
            excluded_rows.append((excel_row_idx, f'ไม่พบแปลง {plot_num} ในระบบ', col_a, col_b, col_c, _cell(row, C['qty']), _cell(row, C['amount']), _cell(row, C['total'])))
            continue

        iso_date, short_date = parse_thai_date(col_a)
        if not iso_date:
            excluded_rows.append((excel_row_idx, 'อ่านวันที่ไม่ได้', col_a, col_b, col_c, _cell(row, C['qty']), _cell(row, C['amount']), _cell(row, C['total'])))
            continue

        # Track new items (not in dictionary)
        if col_c not in item_dict and col_c not in MANUAL_CATEGORIES:
            new_items.add(col_c)

        category = classify_item(col_c, item_dict)
        txn = {
            'date': iso_date,
            'dateText': short_date,
            'name': col_c,
            'qty': round2(_cell(row, C['qty'])),
            'total': round2(_cell(row, C['total'])),
            'category': category,
            'isLabor': category == 'ค่าแรง',
            '_excel_row': excel_row_idx,
        }
        plot_txns[plot_num].append(txn)
        all_input_rows.append((excel_row_idx, plot_num, iso_date, col_c, txn['qty'], txn['total']))

    # ─── Duplicate detection ───
    # Definition: same date + plot + product name + qty + total = potential duplicate
    dup_key_count = defaultdict(list)  # key -> list of excel row indices
    for r in all_input_rows:
        key = (r[1], r[2], r[3], r[4], r[5])  # plot, date, name, qty, total
        dup_key_count[key].append(r[0])
    duplicates = {k: v for k, v in dup_key_count.items() if len(v) > 1}

    txn_count = sum(len(v) for v in plot_txns.values())
    # ─── Safety: ถ้าอ่านไม่ได้เลย ยกเลิก ไม่เขียนทับ index.html (กันข้อมูลกลายเป็นศูนย์) ───
    if txn_count == 0:
        print("   ❌ ไม่พบรายการต้นทุนเลยสักรายการ — ไฟล์อาจผิดรูปแบบ/ผิดชนิด")
        print("      ยกเลิกการทำงาน ไม่เขียน index.html (ป้องกันข้อมูลถดถอยเป็นศูนย์)")
        sys.exit(1)
    print(f"   Captured: {txn_count} transactions across {len(plot_txns)} plots")
    print(f"   Excluded: {len(excluded_rows)} rows (overhead / incomplete / parse error)")
    print(f"   New items needing review: {len(new_items)}")
    print(f"   Duplicate groups: {len(duplicates)}")

    # ─── Build houses array (483) ───
    houses = []
    plot_totals_from_dashboard = {}  # for cross-check
    for plot_num in sorted(plot_metadata.keys()):
        h = dict(plot_metadata[plot_num])  # start with metadata
        items = plot_txns.get(plot_num, [])
        # Strip internal _excel_row field before saving
        for it in items:
            it.pop('_excel_row', None)

        if not items:
            # Inactive plot
            h.update({
                'hasActivity': False,
                'totalCost': 0,
                'materialCost': 0,
                'laborCost': 0,
                'itemCount': 0,
                'byCategory': {},
                'timeline': [],
                'phases': [{'name': pn, 'amount': 0, 'started': False}
                           for pn, _ in PHASE_CATEGORIES],
                'phaseProgress': 0,
                'lastDate': None,
                'items': [],
            })
        else:
            # Active plot — compute aggregates
            items.sort(key=lambda x: x['date'])
            total = round2(sum(i['total'] for i in items))
            labor = round2(sum(i['total'] for i in items if i['isLabor']))
            material = round2(total - labor)

            by_cat = defaultdict(float)
            for it in items:
                by_cat[it['category']] += it['total']
            by_cat = {k: round2(v) for k, v in by_cat.items()}

            # Timeline — group by date, compute cumulative
            by_date = defaultdict(float)
            date_text = {}
            for it in items:
                by_date[it['date']] += it['total']
                date_text[it['date']] = it['dateText']
            timeline = []
            cum = 0
            for d in sorted(by_date.keys()):
                amt = round2(by_date[d])
                cum = round2(cum + amt)
                timeline.append({
                    'date': d,
                    'dateText': date_text[d],
                    'amount': amt,
                    'cumulative': cum,
                })

            # Phases — sum from byCategory
            phases = []
            for phase_name, cat_list in PHASE_CATEGORIES:
                amt = round2(sum(by_cat.get(c, 0) for c in cat_list))
                phases.append({
                    'name': phase_name,
                    'amount': amt,
                    'started': amt > 0,
                })
            phase_progress = round(100 * sum(1 for p in phases if p['started']) / len(phases))

            h.update({
                'hasActivity': True,
                'totalCost': total,
                'materialCost': material,
                'laborCost': labor,
                'itemCount': len(items),
                'byCategory': by_cat,
                'timeline': timeline,
                'phases': phases,
                'phaseProgress': phase_progress,
                'lastDate': items[-1]['dateText'],
                'items': items,
            })
            plot_totals_from_dashboard[plot_num] = total
        houses.append(h)

    # ─── Build full PROJECT_DATA ───
    now = datetime.now()
    project_data = {
        'meta': {
            'project': 'โครงการวันเวลา (One Vela) — อมตะนคร ชลบุรี',
            'generated': now.strftime('%Y-%m-%d %H:%M'),
            'totalPlots': len(houses),
            'currency': 'บาท (รวม VAT)',
        },
        'typeOrder': meta['typeOrder'],
        'phaseNames': meta['phaseNames'],
        'categoryOrder': meta['categoryOrder'],
        'houses': houses,
    }

    # ─── Inject into HTML template ───
    html = TEMPLATE_HTML.read_text(encoding='utf-8')
    # Replace PROJECT_DATA — find by balanced braces
    idx = html.find('const PROJECT_DATA = ')
    start = idx + len('const PROJECT_DATA = ')
    depth, in_str, escape = 0, False, False
    for i in range(start, len(html)):
        c = html[i]
        if escape:
            escape = False; continue
        if c == '\\':
            escape = True; continue
        if c == '"' and not escape:
            in_str = not in_str; continue
        if in_str:
            continue
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    new_json = json.dumps(project_data, ensure_ascii=False, separators=(', ', ': '))
    new_html = html[:start] + new_json + html[end:]

    # ─── Inject Google Sheet sync script (Phase 2) ───
    if SHEET_CONFIG_PATH.exists():
        sheet_cfg = json.loads(SHEET_CONFIG_PATH.read_text(encoding='utf-8'))
        sync_js_template = (TOOLS_DIR / 'sheet_sync.js').read_text(encoding='utf-8')
        sync_js = (sync_js_template
                   .replace('__SHEET_CSV_URL__', sheet_cfg['csv_url'])
                   .replace('__INSTALLMENTS_CSV_URL__', sheet_cfg.get('installments_csv_url', ''))
                   .replace('__SHEET_EDIT_URL__', sheet_cfg.get('edit_url_hint', '')))

        marker = '<!-- SHEET_SYNC -->'
        injected_block = f'\n{marker}\n<script>\n{sync_js}\n</script>\n'

        if marker in new_html:
            # Replace existing injected block (between markers, only one occurrence per page)
            # We use single-marker comment + script; find region [marker .. </script>\n] and replace
            start_m = new_html.find(marker)
            # Find closing </script> after marker
            end_script = new_html.find('</script>', start_m)
            end_block = new_html.find('\n', end_script + len('</script>')) + 1
            new_html = new_html[:start_m] + marker + '\n<script>\n' + sync_js + '\n</script>\n' + new_html[end_block:]
        else:
            # Insert before </body>
            new_html = new_html.replace('</body>', injected_block + '\n</body>')
        print(f"   🔗 Sheet sync injected (CSV URL configured)")
    else:
        print(f"   ⚠ No sheet_config.json found — Dashboard runs in standalone mode")

    OUT_HTML.write_text(new_html, encoding='utf-8')

    # Summary
    active_plots = [h for h in houses if h['hasActivity']]
    total_cost = sum(h['totalCost'] for h in active_plots)
    print(f"\n✅ Output written: {OUT_HTML}")
    print(f"   Active plots: {len(active_plots)}")
    print(f"   Inactive plots: {len(houses) - len(active_plots)}")
    print(f"   Total cost: ฿{total_cost:,.0f}")
    print(f"   Generated at: {project_data['meta']['generated']}")

    # ─── Rebuild deploy zip ───
    deploy_files = ['index.html', 'manifest.json', 'sw.js', 'favicon.png',
                    'icon-192.png', 'icon-192-maskable.png',
                    'icon-512.png', 'icon-512-maskable.png']
    # Include assets folder (photos)
    assets_dir = DASHBOARD_DIR / 'assets'
    asset_files = []
    if assets_dir.exists():
        for p in assets_dir.rglob('*'):
            if p.is_file() and not p.name.startswith('.'):
                asset_files.append(p.relative_to(DASHBOARD_DIR))
    # Write to temp first to avoid permission issues, then copy
    import shutil, tempfile
    tmp_zip = Path(tempfile.gettempdir()) / 'dashboard-deploy-temp.zip'
    if tmp_zip.exists():
        tmp_zip.unlink()
    with ZipFile(tmp_zip, 'w', ZIP_DEFLATED) as zf:
        for fn in deploy_files:
            p = DASHBOARD_DIR / fn
            if p.exists():
                zf.write(p, fn)
        # Include assets folder
        for rel_path in asset_files:
            p = DASHBOARD_DIR / rel_path
            zf.write(p, str(rel_path))
    try:
        if OUT_ZIP.exists():
            OUT_ZIP.unlink()
        shutil.copy(tmp_zip, OUT_ZIP)
    except PermissionError:
        # Fall back: write zip with different name
        alt = DASHBOARD_DIR / f'dashboard-deploy-{now.strftime("%Y%m%d-%H%M")}.zip'
        shutil.copy(tmp_zip, alt)
        print(f"   Deploy zip (alt name): {alt.name}")
        return project_data
    print(f"   Deploy zip: {OUT_ZIP.name} ({OUT_ZIP.stat().st_size:,} bytes)")

    # ─── Build Excel verification report ───
    OUT_REPORT_DIR.mkdir(exist_ok=True)
    report_path = OUT_REPORT_DIR / f'audit_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    # Per-plot comparison: Excel summary sum vs Dashboard sum
    per_plot_excel_total = defaultdict(float)
    for r in rows[2:]:
        col_a = (r[0] or '').strip() if r[0] else ''
        col_b = (r[1] or '').strip() if r[1] else ''
        if 'ทั้งหมด' in col_a:
            m = re.match(r'^(.+?)\s*แปลง\s*(\d+)\s*ทั้งหมด', col_a)
            if m:
                pn = int(m.group(2))
                per_plot_excel_total[pn] += round2(r[6])

    rwb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = rwb.active
    ws1.title = 'สรุปการเทียบ'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill('solid', fgColor='0F2742')
    bold = Font(bold=True)
    green = PatternFill('solid', fgColor='D1FAE5')
    red = PatternFill('solid', fgColor='FEE2E2')
    yellow = PatternFill('solid', fgColor='FEF3C7')

    ws1['A1'] = f'Audit Report — Dashboard ต้นทุนบ้าน วันเวลา'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f'สร้างเมื่อ: {now.strftime("%Y-%m-%d %H:%M")}'
    ws1['A3'] = f'Excel ต้นทาง: {excel_path.name}'

    summary_data = [
        ('', '', ''),
        ('หัวข้อ', 'จำนวน/มูลค่า', 'หมายเหตุ'),
        ('Excel: row ทั้งหมด', len(rows), ''),
        ('Excel: detail rows (ไม่นับ summary)', len(all_input_rows) + len(excluded_rows), ''),
        ('Excel: รวมยอด summary "...ทั้งหมด"', f'฿{excel_total_from_summary:,.2f}', 'ผลรวมจาก row "ทั้งหมด" ของ Excel'),
        ('', '', ''),
        ('Dashboard: บันทึกสำเร็จ', txn_count, f'{len(plot_txns)} แปลง'),
        ('Dashboard: ยอดต้นทุนรวม', f'฿{sum(h["totalCost"] for h in houses):,.2f}', 'รวมทุกแปลง'),
        ('', '', ''),
        ('ถูกข้าม (ดู Sheet 2)', len(excluded_rows), 'overhead / row ที่ไม่ครบ'),
        ('รายการซ้ำ (ดู Sheet 3)', sum(len(v)-1 for v in duplicates.values()), f'{len(duplicates)} กลุ่ม'),
        ('สินค้าใหม่ (ดู Sheet 5)', len(new_items), 'ต้องระบุ category'),
    ]
    for row_idx, (a, b, c) in enumerate(summary_data, start=4):
        ws1.cell(row=row_idx, column=1, value=a)
        ws1.cell(row=row_idx, column=2, value=b)
        ws1.cell(row=row_idx, column=3, value=c)
        if a == 'หัวข้อ':
            for col in (1, 2, 3):
                ws1.cell(row=row_idx, column=col).font = header_font
                ws1.cell(row=row_idx, column=col).fill = header_fill

    # Cross-check (Excel summary vs Dashboard)
    ws1['A17'] = 'ผลเทียบยอด'
    ws1['A17'].font = bold
    diff = excel_total_from_summary - sum(h['totalCost'] for h in houses)
    if abs(diff) < 1:
        ws1['A18'] = '✅ ยอดตรง — บันทึกครบทุก row ของแปลง'
        ws1['A18'].fill = green
    else:
        excluded_sum = sum((r[7] or 0) for r in excluded_rows)
        if abs(diff - excluded_sum) < 1:
            ws1['A18'] = f'⚠ ยอดต่าง ฿{diff:,.2f} = ยอด overhead ที่ข้าม (ดู Sheet 2)'
            ws1['A18'].fill = yellow
        else:
            ws1['A18'] = f'❌ ยอดต่าง ฿{diff:,.2f} (Excel ฿{excel_total_from_summary:,.2f} vs Dashboard ฿{sum(h["totalCost"] for h in houses):,.2f}) — ตรวจสอบ'
            ws1['A18'].fill = red

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 50

    # Sheet 2: Excluded rows
    ws2 = rwb.create_sheet('รายการที่ไม่ได้บันทึก')
    headers = ['Excel row', 'เหตุผล', 'วันที่', 'ลูกค้า/ผู้ขาย', 'สินค้า', 'จำนวน', 'จำนวนเงิน', 'ยอดรวม']
    for i, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
    for r_idx, row in enumerate(excluded_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
    for col_letter, width in zip('ABCDEFGH', [12, 28, 15, 36, 50, 10, 14, 14]):
        ws2.column_dimensions[col_letter].width = width

    # Sheet 3: Duplicates
    ws3 = rwb.create_sheet('รายการซ้ำ')
    ws3.append(['สถานะ', 'แปลง', 'วันที่', 'สินค้า', 'จำนวน', 'ยอดรวม', 'Excel rows ที่ซ้ำกัน', 'กระทบยอด'])
    for c in range(1, 9):
        ws3.cell(row=1, column=c).font = header_font
        ws3.cell(row=1, column=c).fill = header_fill
    for (plot, date, name, qty, total), row_ids in duplicates.items():
        ws3.append([
            f'ซ้ำ {len(row_ids)} ครั้ง',
            plot, date, name, qty, total,
            ', '.join(str(r) for r in row_ids),
            f'฿{(len(row_ids)-1) * total:,.2f} (ส่วนเกิน)',
        ])
        ws3.cell(row=ws3.max_row, column=1).fill = yellow
    if not duplicates:
        ws3.append(['ไม่พบรายการซ้ำ', '', '', '', '', '', '', ''])
        ws3.cell(row=2, column=1).fill = green
    for col_letter, width in zip('ABCDEFGH', [16, 8, 12, 50, 10, 14, 18, 28]):
        ws3.column_dimensions[col_letter].width = width

    # Sheet 4: Per-plot comparison
    ws4 = rwb.create_sheet('เทียบยอดทุกแปลง')
    ws4.append(['แปลง', 'ไทป์', 'Excel ยอดรวม', 'Dashboard ยอดรวม', 'ส่วนต่าง', 'สถานะ'])
    for c in range(1, 7):
        ws4.cell(row=1, column=c).font = header_font
        ws4.cell(row=1, column=c).fill = header_fill
    for plot in sorted(set(list(per_plot_excel_total.keys()) + list(plot_totals_from_dashboard.keys()))):
        excel_t = round2(per_plot_excel_total.get(plot, 0))
        dash_t = round2(plot_totals_from_dashboard.get(plot, 0))
        d = round2(excel_t - dash_t)
        t = plot_metadata.get(plot, {}).get('typeKey', '?')
        status = '✓ ตรง' if abs(d) < 0.5 else f'⚠ ต่าง'
        ws4.append([plot, t, excel_t, dash_t, d, status])
        if abs(d) >= 0.5:
            for c in range(1, 7):
                ws4.cell(row=ws4.max_row, column=c).fill = red
    for col_letter, width in zip('ABCDEF', [10, 8, 18, 18, 14, 12]):
        ws4.column_dimensions[col_letter].width = width

    # Sheet 5: New items needing category review
    ws5 = rwb.create_sheet('สินค้าใหม่_ตรวจ category')
    ws5.append(['ชื่อสินค้า', 'Category ที่ระบบเดา', 'ถูกต้องไหม? (yes/no)', 'Category ที่ถูก'])
    for c in range(1, 5):
        ws5.cell(row=1, column=c).font = header_font
        ws5.cell(row=1, column=c).fill = header_fill
    for item in sorted(new_items):
        cat = classify_item(item, item_dict)
        ws5.append([item, cat, '', ''])
    if not new_items:
        ws5.append(['ไม่มีสินค้าใหม่ — ทุกตัวอยู่ใน dictionary แล้ว', '', '', ''])
        ws5.cell(row=2, column=1).fill = green
    for col_letter, width in zip('ABCD', [50, 26, 18, 26]):
        ws5.column_dimensions[col_letter].width = width

    rwb.save(report_path)
    print(f"   Audit report: {report_path.name} ({report_path.stat().st_size:,} bytes)")

    return project_data


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python3 excel_to_dashboard.py <excel-file>")
        sys.exit(1)
    convert(Path(sys.argv[1]))
